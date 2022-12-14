from openpyxl.styles import colors, PatternFill, Font, Color
from requests.adapters import HTTPAdapter, Retry
from openpyxl import load_workbook
import requests
import shutil
import os

def run():
    INPUT_PATH = 'input/'
    OUTPUT_PATH = 'output/'

    ASSET_TX_URL = {
        'ADA':['https://explorer.cardano.org/en/transaction?id='],
        'ALGO':['https://algoexplorer.io/tx/'],
        'ATOM':['https://atomscan.com/transactions/'],
        'BCH':['https://www.blockchain.com/bch/tx/'],
        'BSV':['https://bsv.tokenview.com/en/tx/'],
        'BTT':['https://bttcscan.com/tx/'],
        'DASH':['https://explorer.dash.org/insight/tx/'],
        'DOGE':['https://dogechain.info/tx/'],
        'DOT':['https://polkascan.io/polkadot/transaction/'],
        'EOS':['https://eosflare.io/tx/'],
        'GAS':['https://neoscan.io/transaction/'],
        'LTC':['https://litecoinblockexplorer.net/tx/'],
        'NANO':['https://nanolooker.com/block/'],
        'NEO':['https://neoscan.io/transaction/'],
        'SOL':['https://explorer.solana.com/tx/'],
        'TRX':['https://tronscan.org/#/transaction/'],
        'VET':['https://explore.vechain.org/transactions/'],
        'XLM':['https://stellarchain.io/tx/'],
        'XRP':['https://xrpscan.com/tx/'],
        'ZEC':['https://zcash.tokenview.com/en/tx/'],
        'BTC':['https://www.blockchain.com/btc/tx/']
    }

    SCAN_TX_URL = ['https://etherscan.io/tx/', 'https://polygonscan.com/tx/', 'https://bscscan.com/tx/', 'https://snowtrace.io/tx/', 'https://ftmscan.com/tx/']

    def verify_tx_url(url):
        HEADERS = {
            "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/74.0.3729.169 Safari/537.36" ,
            'referer':'https://www.google.com/'
        }
        response = requests.Session()
        retries = Retry(total=15, backoff_factor=.1)
        response.mount('https://', HTTPAdapter(max_retries=retries))
        result = response.get(url, headers=HEADERS, timeout=2)
        if result.status_code != 200 or 'unable to locate this TxnHash' in result.text:
            return False
            
        return True
    
    _, _, files = next(os.walk(INPUT_PATH))

    for file in files:
        shutil.copy(f'{INPUT_PATH}{file}', OUTPUT_PATH)
        work_book = load_workbook(f'{OUTPUT_PATH}{file}', data_only=True)

        for worksheet_idx in range(len(work_book.worksheets)):
            tx_id_idx = None
            tx_url_idx = None
            asset_idx = None
            work_book.active = worksheet_idx
            work_sheet = work_book.active
            print(f'Working with {file} - {work_sheet.title}')
            
            def invalidate(row, col):
                work_sheet[row][col].value = 'Null'
                work_sheet[row][col].fill = PatternFill(start_color='F2D3D7', end_color='F2D3D7', fill_type = 'solid')
                work_sheet[row][col].font = Font(color='9C0039')            
                
            def validate(row, col, value):
                work_sheet[row][col].value = value
                work_sheet[row][col].fill = PatternFill(start_color='C3ECCB', end_color='C3ECCB', fill_type = 'solid')
                work_sheet[row][col].font = Font(color='006100')             
           
            for column_idx, column_name in enumerate(work_sheet[1]):
                if column_name.value in ('Transaction Details', 'Transaction ID', 'Transaction Detail'):
                    tx_id_idx = column_idx
                elif column_name.value == 'Blockchain URL':
                    tx_url_idx = column_idx
                elif column_name.value in ('Asset', 'Currency'):
                    asset_idx = column_idx

            row_count = work_sheet.max_row
           
            for row_idx in range(2, row_count + 1):
                print(f'Working with row #{row_idx}/{row_count}')
                if tx_id_idx is not None and tx_url_idx is not None:
                    tx_id = work_sheet[row_idx][tx_id_idx].value
                    
                    if not tx_id:
                        invalidate(row_idx, tx_url_idx)
                        continue
                        
                    prefixes = []
                    if asset_idx is not None:
                        asset = work_sheet[row_idx][asset_idx].value
                        if asset is None:
                            prefixes = SCAN_TX_URL
                        else:
                            prefixes = ASSET_TX_URL.get(asset.upper(), SCAN_TX_URL)
                            
                    for prefix in prefixes:
                        tx_url = f'{prefix}{tx_id}'
                        if prefix in SCAN_TX_URL and not tx_id.startswith('0x'):
                            tx_url = f'{prefix}0x{tx_id}'
                            
                        try:
                            if verify_tx_url(tx_url):
                                validate(row_idx, tx_url_idx, f'=HYPERLINK("{tx_url}", "Verified")')
                                break
                            else:
                                invalidate(row_idx, tx_url_idx)
                        except Exception as e:
                            invalidate(row_idx, tx_url_idx)
                    
        print('Verification Complete.')
        work_book.save(f'{OUTPUT_PATH}{file}')
    
